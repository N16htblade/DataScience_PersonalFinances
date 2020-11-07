import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import os, sys
import subprocess
import csv
import numpy as np

inputFiles = [file for file in os.listdir('.\Budget\Input') if file.endswith('.csv')] #Check for input files (ignore non .csv), concate into main datafile and remove duplicated entries.
print (f'Found {len(inputFiles)} files to process.')

allYearData = pd.DataFrame()

for file in inputFiles:
    df = pd.read_csv('./Budget/Input/' + file)
    allYearData = pd.concat([allYearData, df])

allYearData = allYearData.drop_duplicates(keep='last')
allYearData.to_csv('./Budget/Output/allYearData.csv', index=False)

count_row = allYearData.shape[0]
if count_row <1:
    print ('No data available to process, please make sure raw data is available in the Input folder.')
    exit
else:
    print ('Data loaded successfully and duplicated entries removed.\nProcessing data, please wait...')


df = pd.read_csv('./Budget/Output/allYearData.csv') #Load the main data file.
df.rename(columns = {'Transaction Date':'TransactionDate', 'Description 1':'Description1', 'Description 2':'Description2'}, inplace = True) #Remove spaces in column name.
df = df.loc[~df['Description1'].str.contains('Transfer') & ~df['Description1'].str.contains('MERCI') & ~df['Description1'].str.contains('ONLINE') & ~df['Description1'].str.contains('PAYBACK')]#Remove exchanges between accounts.

with open (os.path.join(sys.path[0], "Calendar.csv")) as csv_file: #Transform transaction date into Month and sort based on Calendar.
    calendar_reader = csv.reader(csv_file, delimiter=",")
    months = {int(rows[0]):rows[1] for rows in calendar_reader}

df['TransactionDate'] = pd.to_datetime(df['TransactionDate']) 
df['TransactionDate'] = df['TransactionDate'].dt.month
df = df.sort_values('TransactionDate').reset_index()
df['Month'] = df.TransactionDate.map(months)


with open (os.path.join(sys.path[0], "Category.csv")) as csv_file:#Simplify description and map to category based on Dictionary.
    category_reader = csv.reader(csv_file, delimiter=",")
    categoryByDescription = {rows[0]:rows[1] for rows in category_reader}

df["Description"] = df.Description1.str.split().str.get(0) 
tempdf=df.loc[(df['Description'].str.len() < 3) | (df['Description'] == 'THE')] #Check for short or common denominator Description and add second string to avoid issues.
for index, row in tempdf.iterrows():
    q = row.Description1
    q = q.split()[:2]
    q = q[0] + ' ' + q[1]
    df.loc[index, 'Description'] = q


mainCategories = {'Housing' : 'Rent, Insurance, Other.', #Check for transactions without a Category mapped, verbose.
                  'Food': 'Groceries, Take-Out, Coffee, Other.',
                  'Transportation': 'Auto, AutoInsurance, Fuel, Other.',
                  'Personal': 'Gym, Hair, PersonalCare, Other.',
                  'Banking': 'Investing, Fee, Taxes, Other.',
                  'Entertainment': 'Netflix, Games, Movies, Other.',
                  'Utilities': 'Power, Internet, Phone, Other.'}
descriptionList = list(set(df["Description"]))
for i in descriptionList:
    if i in categoryByDescription.keys():
        pass
    else:
        print (f'\nFollowing transactions does not have a category assigned: {i}')
        print (f'Please enter Main and Secondary category for {i} vendor.')
        value1 = input(f'Main category (Housing, Food, Transportation, Personal, Banking, Entertainment or Utilities):  ')
        if value1 in mainCategories.keys():
            options = mainCategories[f'{value1}']
            value2 = input(f'Sub Category (Ex: {options}): ')
            #value1 = input(f'Incorrect category, please use one of the listed categories (Housing, Food, Transportation, Personal, Banking, Entertainment or Utilities):  ')
        newValue = f'{value1} {value2}'
        categoryByDescription[i] = newValue


with open (os.path.join(sys.path[0], "Category.csv"), 'w', newline='') as f: #update Category dictionary
    w = csv.writer(f)
    w.writerows(categoryByDescription.items())


pat = '|'.join(r"\b{}\b".format(x) for x in categoryByDescription.keys()) #Process Descriptions and map to Categories.
df['CategoryAll'] = df['Description'].str.extract('('+ pat + ')', expand=False).map(categoryByDescription)
df['Category1'] = df.CategoryAll.str.split().str.get(0)
df['Category2'] = df.CategoryAll.str.split().str.get(1)
df['CAD$'] = df['CAD$'].astype(int)
df = df.drop(['CategoryAll', 'Account Number', 'TransactionDate', 'Account Type', 'USD$', 'Cheque Number', 'Description1', 'Description2', 'index'], axis = 1) #Remove unnecesary/empty columns.
df = df.groupby(['Month', 'Category1', 'Category2', 'Description'], sort=False).sum().reset_index() #Sort and merge database on Month, Categories and Description. Save to output file.
df.to_excel('./Budget/Output/finalData.xlsx', sheet_name='Data', index=False)
wb = openpyxl.load_workbook('./Budget/Output/finalData.xlsx')
ws = wb.create_sheet('Graphs',0)


monthlyIncome = df[df['Category1'] == 'Income'] #Create top plot for Yearly revenue based on monthly Income/Expenses.
monthlyIncome = monthlyIncome.groupby(['Month'], sort=False).sum().reset_index()
monthlyExpenses = df[df['Category1'] != 'Income']
monthlyExpenses = monthlyExpenses.groupby(['Month'], sort=False).sum().reset_index()
monthlyExpenses['CAD$'] = monthlyExpenses['CAD$'].abs()
plt.figure(figsize=(15.6,2))
plt.title('Monthly Income / Expenses')
plt.plot(monthlyIncome['Month'], monthlyIncome['CAD$'], label='Income', color='g', marker='.', markersize=10, markeredgecolor='k', markeredgewidth=0.5)
plt.plot(monthlyExpenses['Month'], monthlyExpenses['CAD$'], label='Expenses', color='r', marker='.', markersize=10, markeredgecolor='k', markeredgewidth=0.5)
plt.legend()
plt.savefig('./Budget/Temp/tempI.png',dpi=100, bbox_inches='tight')
img = openpyxl.drawing.image.Image('./Budget/Temp/tempI.png')
img.anchor = 'A1'
ws.add_image(img)
wb.save('./Budget/Output/finalData.xlsx')
plt.close()


monthly = pd.DataFrame() #Following section generates data and the Monthly bar plot.
monthly['Month'] = ('Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sept', 'Oct', 'Nov', 'Dec')

def monthlyBars (category, monthly):
    monthlyx = df[df['Category1'] == f'{category}']
    monthlyx = monthlyx.groupby(['Month'], sort=False).sum().reset_index()
    cols = ['Month']
    monthly = monthly.merge(monthlyx, how='left')
    monthly.fillna(0, inplace=True)
    monthly['CAD$'] = monthly['CAD$'].abs()
    monthlyExpence = list(monthly['CAD$'])
    return monthlyExpence

housing = monthlyBars ('Housing', monthly)
utilities = monthlyBars ('Utilities', monthly)
food = monthlyBars ('Food', monthly)
transportation = monthlyBars ('Transportation', monthly)
personal = monthlyBars ('Personal', monthly)
banking = monthlyBars ('Banking', monthly)
entertainment = monthlyBars ('Entertainment', monthly)
dftemp = pd.DataFrame({'Housing' : housing,'Utilities' : utilities, 'Food' : food,'Transportation' : transportation,'Personal' : personal,'Banking' : banking,'Entertainment' : entertainment})

fig, ax = plt.subplots()
fig.set_size_inches(7,9)
dftemp.plot.barh(stacked=True, ax=ax)
ax.set_title("Monthly Expenses by Group")
ax.legend(loc='lower right')
ax.set_yticklabels(monthly['Month'])
plt.savefig('./Budget/Temp/tempIEM.png',dpi=100, bbox_inches='tight')
img = openpyxl.drawing.image.Image('./Budget/Temp/tempIEM.png')
img.anchor = 'U1'
ws.add_image(img)
wb.save('./Budget/Output/finalData.xlsx')
plt.close()

pieLabels = []
pieValues = []

def yearCalc (): #Function to create main Total spendings donut chart;
    yearExpenses = abs(df[df['Category1'] != 'Income']['CAD$'].sum())
    limit = yearExpenses * 0.05
    categoryOther = 0
    tempCategoriesContainer = {}

    for i in mainCategories.keys():
        categoryExpenses = abs(df[df['Category1'] == f'{i}']['CAD$'].sum())
        if categoryExpenses < limit:
            categoryOther += categoryExpenses
        else:
            tempCategoriesContainer[f'{i}'] = categoryExpenses

    if categoryOther > 0:
        tempCategoriesContainer['Other'] = categoryOther

    tempCategoriesContainer = {k: v for k, v in sorted(tempCategoriesContainer.items(), key=lambda item: item[1], reverse=True)}

    pieLabelsTemp = list(tempCategoriesContainer)
    for i in pieLabelsTemp:
        pieLabels.append(i)
    
    pieValuesTemp = list(tempCategoriesContainer.values())
    for i in pieValuesTemp:
        pieValues.append(i)

    donutChart (pieLabels, pieValues, 'Main', 'Total Expenses', 'A12', yearExpenses)


def yearlyCategoryCalc (category, title, location, total): #Generate donutcharts based on category, name and location
    yearly = df[df['Category1'] == f'{category}']
    yearly = yearly.groupby(['Category2']).sum().reset_index()
    yearly['CAD$'] = yearly['CAD$'].abs()
    yearly = yearly.sort_values(by='CAD$', ascending=False)
    pieLabels = yearly.Category2.values.tolist()
    pieValues = yearly['CAD$'].values.tolist()
    donutChart (pieLabels, pieValues, category, title, location, total)


def donutChart (pieLabels, pieValues, category, title, location, total): #Donutchart function;
    fig, ax1 = plt.subplots()
    explodeTemp = []
    for i in pieLabels:
        explodeTemp.append(0.02)
    explode = tuple(explodeTemp)
    colors = ['#33bbff', '#F65314' ,'#FBBB00', '#7CBB00', '#9999ff', '#ff99ff', '#ffcc99']
    plt.pie(pieValues, labels=pieLabels, colors=colors, autopct=lambda p : '{:.1f}%\n${:,.0f}'.format(p,p * sum(pieValues)/100), pctdistance=0.80, explode=explode, startangle=90, counterclock=False)
    centre_circle = plt.Circle((0,0),0.60,fc='white')
    label = ax1.annotate(f'{title}\n${total}', fontsize = 12, xy=(0, -0.10), ha="center")
    fig = plt.gcf()
    fig.gca().add_artist(centre_circle)
    plt.savefig(f'./Budget/Temp/{category}.png',dpi=70, bbox_inches='tight')
    img = openpyxl.drawing.image.Image(f'./Budget/Temp/{category}.png')
    img.anchor = f'{location}'
    ws.add_image(img)
    wb.save('./Budget/Output/finalData.xlsx')
    plt.close()

#Find a way to get rid of this stuff in functions above
yearIncome = df[df['Category1'] == 'Income']['CAD$'].sum() 
yearExpenses = abs(df[df['Category1'] != 'Income']['CAD$'].sum())
yearExpensesFood = abs(df[df['Category1'] == 'Food']['CAD$'].sum())
yearExpensesPersonal = abs(df[df['Category1'] == 'Personal']['CAD$'].sum())
yearExpensesTransportation = abs(df[df['Category1'] == 'Transportation']['CAD$'].sum())
yearExpensesUtilities = abs(df[df['Category1'] == 'Utilities']['CAD$'].sum())
yearExpensesHousing = abs(df[df['Category1'] == 'Housing']['CAD$'].sum())
yearExpensesEntertainment = abs(df[df['Category1'] == 'Entertainment']['CAD$'].sum())
yearExpensesBanking = abs(df[df['Category1'] == 'Banking']['CAD$'].sum())

yearCalc() #Create main pie chart based on total expenses and main category.
yearlyCategoryCalc('Housing', 'Housing', 'F12', yearExpensesHousing)
yearlyCategoryCalc('Food', 'Food', 'K12', yearExpensesFood)
yearlyCategoryCalc('Transportation', 'Transportation', 'P12', yearExpensesTransportation)
yearlyCategoryCalc('Personal', 'Personal', 'A26', yearExpensesPersonal)
yearlyCategoryCalc('Banking', 'Banking', 'F26', yearExpensesBanking)
yearlyCategoryCalc('Entertainment', 'Entertainment', 'K26', yearExpensesEntertainment)
yearlyCategoryCalc('Utilities', 'Utilities', 'P26', yearExpensesUtilities)
ws = wb.create_sheet('Main',0)

ytdRemaining = yearIncome - yearExpenses
pieLabels = ('Income', 'Expenses')
pieValues = (yearIncome, yearExpenses)
donutChart (pieLabels, pieValues, 'tempYTD', 'YTD Balance', 'A1', ytdRemaining)

img = openpyxl.drawing.image.Image('./Budget/Temp/main.png')
img.anchor = 'Q1'
ws.add_image(img)
wb.save('./Budget/Output/finalData.xlsx')
plt.close()


plain12Text = openpyxl.styles.Font(size=12) #xls formating - styles
i11GreyText = openpyxl.styles.Font(italic=True, color='666666', size=11)
b11Text = openpyxl.styles.Font(bold=True, size=11) 
b12Text = openpyxl.styles.Font(bold=True, size=12)
bU12Text = openpyxl.styles.Font(bold=True, underline='single', size=12)
plainNumber = openpyxl.styles.Font(size=11)
bPosNum = openpyxl.styles.Font(bold=True, color='7CBB00', size=11)
#bUPosNum = openpyxl.styles.Font(bold=True, underline='single', color='7CBB00', size=12)
bUPosNum = openpyxl.styles.Font(bold=True, color='7CBB00', size=12)
bNegNum = openpyxl.styles.Font(bold=True, color='F65314', size=11)
#bUNegNum = openpyxl.styles.Font(bold=True, underline='single', color='F65314', size=12)
bUNegNum = openpyxl.styles.Font(bold=True, color='F65314', size=12)
rightAlign = openpyxl.styles.Alignment(horizontal='right')
centerAlign = openpyxl.styles.Alignment(horizontal='center')

def addPlainText (cellText, cellLocation):  #xls formating - style based functions
    ws[f'{cellLocation}'] = cellText
    ws[f'{cellLocation}'].font = plain12Text

def addItalicText (cellText, cellLocation):
    ws[f'{cellLocation}'] = cellText
    ws[f'{cellLocation}'].font = i11GreyText

def addMinorText (cellText, cellLocation):
    ws[f'{cellLocation}'] = cellText
    ws[f'{cellLocation}'].font = b11Text

def addMajorText (cellText, cellLocation):
    ws[f'{cellLocation}'] = cellText
    ws[f'{cellLocation}'].font = b12Text

def addMajorUText (cellText, cellLocation):
    ws[f'{cellLocation}'] = cellText
    ws[f'{cellLocation}'].font = bU12Text

def addPlainNumber (cellValue, cellLocation): 
    ws[f'{cellLocation}'] = f'{cellValue}$'
    ws[f'{cellLocation}'].font = plainNumber
    ws[f'{cellLocation}'].alignment = centerAlign

def addPlainNumberPct (cellValue, cellLocation):
    ws[f'{cellLocation}'] = f'{cellValue}%'
    ws[f'{cellLocation}'].font = plainNumber
    ws[f'{cellLocation}'].alignment = centerAlign

def addMinorNumber (cellValue, cellLocation):
    ws[f'{cellLocation}'] = f'{cellValue}$'
    if cellValue > 0:
        ws[f'{cellLocation}'].font = bPosNum
        ws[f'{cellLocation}'].alignment = rightAlign
    else:
        ws[f'{cellLocation}'].font = bNegNum
        ws[f'{cellLocation}'].alignment = rightAlign

def addMajorNumber (cellValue, cellLocation):
    ws[f'{cellLocation}'] = f'{cellValue}$'
    if cellValue > 0:
        ws[f'{cellLocation}'].font = bUPosNum
        ws[f'{cellLocation}'].alignment = rightAlign
    else:
        ws[f'{cellLocation}'].font = bUNegNum
        ws[f'{cellLocation}'].alignment = rightAlign

def addTableHeadText (cellText, cellLocation):
    ws[f'{cellLocation}'] = cellText
    ws[f'{cellLocation}'].font = b11Text
    ws[f'{cellLocation}'].alignment = centerAlign

ws.column_dimensions['G'].width = 13    #xls formating - column row adjustments
ws.column_dimensions['H'].width = 7
ws.column_dimensions['K'].width = 15 
ws.column_dimensions['L'].width = 14
ws.column_dimensions['M'].width = 10
ws.column_dimensions['N'].width = 10
ws.column_dimensions['O'].width = 16

addTableHeadText ('Month Avg.', 'L17')
addTableHeadText ('YTD Total', 'M17')
addTableHeadText ('YTD %', 'N17')
addTableHeadText ('Recommended', 'O17')

graphLabels = []    #list initiation for bar graph - Current vs Recomm;
realLevel = []
recommendedLevel = []

def categoryTable (category, ytdTotal, dataRowY, recommendedPerc, textRowY):  #Generates Monthly / Recommended table, feeds parameters for bar graph;
    noMonths = df['Month'].nunique()
    monthAvg = ytdTotal / noMonths
    monthAvg = int(monthAvg)
    ytdPercent = int(100 * ytdTotal / yearExpenses)

    addMinorText (f'{category}', f'K{dataRowY}')
    addPlainNumber (monthAvg, f'L{dataRowY}')
    addPlainNumber (ytdTotal, f'M{dataRowY}')
    addPlainNumberPct (ytdPercent, f'N{dataRowY}')
    addPlainNumberPct (recommendedPerc, f'O{dataRowY}')

    realLevel.append(ytdPercent)
    graphLabels.append(category)
    recommendedLevel.append(recommendedPerc)

    tooHigh = recommendedPerc * 1.25
    high = recommendedPerc * 1.08
    low = recommendedPerc * 0.92
    tooLow = recommendedPerc * 0.75

    if ytdPercent > tooHigh:
        printable = f'    {category} expenses are at a Very High level, please consider reviewing this category further.'
    elif ytdPercent > high:
        printable = f'    {category} expenses are Withing Acceptable limits, while on the Higher End of recommended level.'
    elif ytdPercent > low:
        printable = f'    {category} expenses are at recommended levels. Keep them here or lower if possible.'
    elif ytdPercent > tooLow:
        printable = f'    {category} expenses are on the Lower End of recommended levels. Good work.'
    else:
        printable = f'    {category} expenses are considerably bellow recommended levels. While not necesarly a bad thing, a review of this category is recommended.'
    
    addItalicText (printable, f'B{textRowY}')


categoryTable('Housing', yearExpensesHousing, 18, 30, 29)
categoryTable('Transportation', yearExpensesTransportation, 19, 17, 30)
categoryTable('Banking', yearExpensesBanking, 20, 15, 31)
categoryTable('Food', yearExpensesFood, 21, 15, 32)
categoryTable('Personal', yearExpensesPersonal, 22, 13, 33)
categoryTable('Entertainment', yearExpensesEntertainment, 23, 5, 34)
categoryTable('Utilities', yearExpensesUtilities, 24, 5, 35)

x = np.arange(len(graphLabels)) #Initiate and control bar graph;
width = 0.4
fig, ax = plt.subplots()
fig.set_size_inches(9,3)
rects1 = ax.bar(x - width/2, realLevel, width, color='#CBF1F5', label='Current')
rects2 = ax.bar(x + width/2, recommendedLevel, width, color='#71C9CE', label='Recommended')
ax.set_title('Current Expenses vs Recommended')
#ax.set_ylabel('Percent (%)')
ax.set_xticks(x)
ax.set_xticklabels(graphLabels)
ax.legend()

def autolabel(rects):
    """Attach a text label above each bar in *rects*, displaying its height."""
    for rect in rects:
        height = rect.get_height() 
        if height < 10:
            ax.annotate('{:.0f}%'.format(height),
                        xy=(rect.get_x() + rect.get_width() / 2, height),
                        xytext=(0, 3),
                        textcoords="offset points",
                        ha='center', va='bottom')
        else:
            ax.annotate('{:.0f}%'.format(height),
                        xy=(rect.get_x() + rect.get_width() / 2, height/2),
                        xytext=(0, 3),
                        textcoords="offset points",
                        ha='center', va='bottom')            

autolabel(rects1)
autolabel(rects2)

plt.savefig('./Budget/Temp/tempLvl.png', bbox_inches='tight')
img = openpyxl.drawing.image.Image('./Budget/Temp/tempLvl.png')
img.anchor = 'G1'
ws.add_image(img)
wb.save('./Budget/Output/finalData.xlsx')
plt.close()

addMajorUText ('YTD Account Balance:', 'B17')
yearInvestements = abs(df[df['Category2'] == 'Investing']['CAD$'].sum())    #Display totals in green/red text
addMajorText ('Income:', 'B18')
addMajorNumber (yearIncome, 'D18')
addMajorText ('Expenses:', 'B19')
onlyExpenses = yearExpenses - yearInvestements
investedpct = int(100 * yearInvestements / yearExpenses)
onlyExpensespct = int(100 * onlyExpenses / yearExpenses)
onlyExpenses = -onlyExpenses
yearExpenses = -yearExpenses
addMajorNumber (yearExpenses, 'D19')
addPlainText ('--------------------------------------', 'B20')
addMajorText ('Balance:', 'B21')
addMajorNumber (ytdRemaining, 'D21')

addMajorUText ('YTD Expenses Overview:', 'G17')
addMajorText ('Investments:', 'G18')
addMajorNumber (yearInvestements, 'I18')
addMajorText ('Expenses:', 'G19')
addMajorNumber (onlyExpenses, 'I19')
addPlainNumberPct (investedpct, 'H18')
addPlainNumberPct (onlyExpensespct, 'H19')

thin = openpyxl.styles.Side(border_style='thin', color='FF000000')  #Initiates border for text area;

def set_border (ws, side, cell_range):
    rows = ws[cell_range]
    if side == "left":
        for row in rows:
            row[0].border = openpyxl.styles.Border(left=thin)
    elif side == 'right':
        for row in rows:
            row[-1].border = openpyxl.styles.Border(right=thin)
    elif side == 'top':
        for row in rows:
            for c in rows[0]:
                c.border = openpyxl.styles.Border(top=thin)
    else:
        for row in rows:
            for c in rows[-1]:
                c.border = openpyxl.styles.Border(bottom=thin)

set_border(ws, 'left', 'B29:B35')
set_border(ws, 'right', 'O29:O35')
set_border(ws, 'top', 'C28:N28')
set_border(ws, 'bottom', 'C36:N36')
top_left_corner = ws['B28']
top_left_corner.border = openpyxl.styles.Border(top=thin, left=thin)
top_right_corner = ws['O28']
top_right_corner.border = openpyxl.styles.Border(top=thin, right=thin)
bottom_left_corner = ws['B36']
bottom_left_corner.border = openpyxl.styles.Border(bottom=thin, left=thin)
bottom_right_corner = ws['O36']
bottom_right_corner.border = openpyxl.styles.Border(bottom=thin, right=thin)

for s in range(len(wb.sheetnames)): #change existing sheets background to white;
    wb.active = s
    ws = wb.active
    for rows in ws.iter_rows(min_col=1, max_col=30, min_row=1, max_row=39):
        for cell in rows:
            cell.fill = openpyxl.styles.PatternFill(start_color='00FFFFFF', end_color='00FFFFFF', fill_type = "solid")
wb.active = 0
wb.save('./Budget/Output/finalData.xlsx')

finalDataLocation = os.getcwd() + '/Budget/Output/finalData.xlsx'   #save final document with all data and charts

openFile = input('Would you like to open the Final Data file? (Yes / No): ')
if openFile == 'y':
    subprocess.Popen([finalDataLocation], shell=True)
else:
    pass